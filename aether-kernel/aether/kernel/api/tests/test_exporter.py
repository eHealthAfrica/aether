# Copyright (C) 2023 by eHealth Africa : http://www.eHealthAfrica.org
#
# See the NOTICE file distributed with this work for additional information
# regarding copyright ownership.
#
# Licensed under the Apache License, Version 2.0 (the "License");
# you may not use this file except in compliance with
# the License.  You may obtain a copy of the License at
#
#   http://www.apache.org/licenses/LICENSE-2.0
#
# Unless required by applicable law or agreed to in writing,
# software distributed under the License is distributed on an
# "AS IS" BASIS, WITHOUT WARRANTIES OR CONDITIONS OF ANY
# KIND, either express or implied.  See the License for the
# specific language governing permissions and limitations
# under the License.

import json
from unittest import mock
import os
import requests
import tempfile
import zipfile

from copy import deepcopy
from random import shuffle

from openpyxl import load_workbook

from django.contrib.auth import get_user_model
from django.core.files.uploadedfile import SimpleUploadedFile
from django.db import connection
from django.db.models import F
from django.test import TransactionTestCase, TestCase, override_settings, tag
from django.urls import reverse

from aether.kernel.api import models

from aether.kernel.api.entity_extractor import run_extraction
from aether.kernel.api.project_artefacts import upsert_project_with_avro_schemas
from aether.kernel.api.utils import safe_sleep

from aether.kernel.api.exporter import (
    __filter_paths as filter_paths,
    __filter_headers as filter_headers,
    __order_headers as order_headers,
    __flatten_dict as flatten_dict,
    __get_label as get_label,

    __generate_csv_files as gen_csv_files,
    __prepare_xlsx as gen_xlsx,
    __prepare_zip as gen_zip,

    execute_records_task,
    execute_attachments_task,

    CSV_FORMAT,
    XLSX_FORMAT,
    MAX_SIZE,
    DEFAULT_OPTIONS,
)


here = os.path.dirname(os.path.realpath(__file__))


EXAMPLE_PATHS = [
    'country',
    'region',
    'name',
    'location',
    'location.latitude',
    'location.longitude',
    'location.altitude',
    'location.accuracy',
    'location_none',
    'location_none.latitude',
    'location_none.longitude',
    'location_none.altitude',
    'location_none.accuracy',
    'image',
    'number',
    'number2',
    'date',
    'datetime',
    'option',
    'option_a',
    'option_a.choice_a',
    'option_b',
    'option_b.choice_b',
    'lang',
    'lang.#',
    'iterate',
    'iterate.#',
    'iterate.#.index',
    'iterate.#.value',
    'iterate_one',
    'iterate_one.#',
    'iterate_one.#.item',
    'iterate_none',
    'iterate_none.#',
    'iterate_none.#.nothing',
    'id',
]

EXAMPLE_LABELS = {
    '_id': 'xForm ID',
    '_version': 'xForm version',
    'country': 'Country',
    'region': 'Region',
    'name': 'What is your name?',
    'location': 'Collect your GPS coordinates',
    'location.latitude': 'latitude',
    'location.longitude': 'longitude',
    'location.altitude': 'altitude',
    'location.accuracy': 'accuracy',
    'location_none': 'Ignore your GPS coordinates',
    'location_none.latitude': 'latitude',
    'location_none.longitude': 'longitude',
    'location_none.altitude': 'altitude',
    'location_none.accuracy': 'accuracy',
    'image': 'Take a picture',
    'number': 'How many?',
    'number2': 'Percentage',
    'date': 'When?',
    'datetime': 'At?',
    'option': 'Choice (A/B)',
    'option_a': 'Option A',
    'option_a.choice_a': 'Choice A',
    'option_b': 'Option B',
    'option_b.choice_b': 'Choice B',
    'lang': 'Spoken languages',
    'iterate': 'Indicate loop elements',
    'iterate.#.index': 'Index',
    'iterate.#.value': 'Value',
    'iterate_one': 'Indicate one',
    'iterate_one.#.item': 'Item',
    'iterate_none': 'Indicate none',
    'iterate_none.#.nothing': 'None',
    'id': 'ID',
}


def helper__generate_file(
    temp_dir,
    data,
    paths=[],
    labels={},
    file_format=CSV_FORMAT,
    filename='export',
    offset=0,
    limit=MAX_SIZE,
    options=DEFAULT_OPTIONS,
):
    '''
    Generates an XLSX/ZIP (of CSV files) file with the given data.

    - ``data`` a queryset with two main properties ``EXPORT_FIELD_ID``
        and ``EXPORT_FIELD_DATA``.
    - ``paths`` is a list with the allowed jsonpaths.
    - ``labels`` is a dictionary whose keys are the jsonpaths
      and the values the linked labels to use as header for that jsonpath.
    - ``file_format``, expected values ``xlsx`` or ``csv``.
    - ``options`` the export options.
    '''

    sql, params = data.query.sql_with_params()
    with connection.cursor() as cursor:
        sql_sentence = cursor.mogrify(sql, params).decode('utf-8')

    csv_files = gen_csv_files(temp_dir,
                              sql_sentence,
                              paths,
                              labels,
                              offset,
                              limit,
                              options,
                              )

    if file_format == XLSX_FORMAT:
        return gen_xlsx(temp_dir, csv_files, filename)
    else:
        return gen_zip(temp_dir, csv_files, filename)


class ExporterTest(TestCase):

    def test__flatten_dict(self):
        item = {
            'a': {
                'b': 1,
                'z': 'z',
            },
            'c': {
                'd': [{'f': 2}],
            },
            'e': [1, 2, 3],
        }
        expected = {
            'a.b': 1,
            'a.z': 'z',
            'c.d': [{'f': 2}],
            'e': [1, 2, 3],
        }
        expected_flatten = {
            'a.b': 1,
            'a.z': 'z',
            'c.d.1.f': 2,
            'e.1': 1,
            'e.2': 2,
            'e.3': 3,
        }

        self.assertEqual(flatten_dict({}), {})
        self.assertEqual(flatten_dict(item), expected)
        self.assertEqual(flatten_dict(item, flatten_list=True), expected_flatten)
        self.assertEqual(flatten_dict(flatten_dict(item)), expected)  # idempotent

    def test__filter_paths(self):
        paths = [
            'a',
            'a.b',
            'a.b.*',
            'a.b.*.#',
            'a.b.*.#.x',
            'a.c',
            'a.c.#',
            'a.c.#.y',
            'a.d',
            'a.d.?',
            'a.d.?.e',
            'a.f',
            'a.f.g',
            'z',
        ]
        expected = [
            'a.b',
            'a.c',
            'a.d',
            'a.f.g',
            'z',
        ]

        self.assertEqual(filter_paths(paths), expected)
        self.assertEqual(filter_paths(filter_paths(paths)), expected)

    def test__filter_headers(self):
        prefix = ['@', '@id']
        headers = ['a', 'x', 'z', 'c', 'd']
        # nothing changes
        self.assertEqual(filter_headers([], '$', headers), headers)

        # includes prefix, filters and orders the headers
        self.assertEqual(filter_headers(['a', 'w', 'd', 'z'], '$', headers), prefix + ['a', 'd', 'z'])

    def test__filter_headers__list(self):
        paths = ['b', 'a']  # not in alphabetical order
        prefix = ['@', '@id']

        expected = [
            'b.1', 'b.2', 'b.3', 'b.4', 'b.5',
            'a.1', 'a.2', 'a.3', 'a.4', 'a.5',
        ]
        headers = deepcopy(expected)

        for _ in range(5):
            shuffle(headers)  # change the order of the elements
            self.assertNotEqual(headers, expected)
            self.assertEqual(filter_headers(paths, '$', headers), prefix + expected)

    def test__filter_headers__nested_list(self):
        paths = ['b', 'a']  # not in alphabetical order
        prefix = ['@', '@id']

        expected = [
            'b.1.1', 'b.1.2', 'b.1.3', 'b.1.4', 'b.1.5', 'b.2.1',
            'a.1.1', 'a.1.2', 'a.1.3', 'a.2.1', 'a.2.2', 'a.3.1',
        ]
        headers = deepcopy(expected)

        for _ in range(5):
            shuffle(headers)  # change the order of the elements
            self.assertNotEqual(headers, expected)
            self.assertEqual(filter_headers(paths, '$', headers), prefix + expected)

    def test__order_headers__documented_case(self):
        headers = [
            'ZZZ',
            'w.2.b.1',
            'w.1.a.1',
            'w.2.a',
            'XXX',
            'b.2',
            'w.3',
            'w.2.b.2',
            'YYY',
            'c.1',
            'w.1.c.1',
            'w.1.c.2',
            'c.2',
            'b.4',
        ]
        expected = [
            'ZZZ',
            'w.1.a.1',
            'w.1.c.1',
            'w.1.c.2',
            'w.2.b.1',
            'w.2.b.2',
            'w.2.a',
            'w.3',
            'XXX',
            'b.2',
            'b.4',
            'YYY',
            'c.1',
            'c.2',
        ]

        self.assertEqual(order_headers(headers), expected)

    def test__get_label(self):
        labels = {
            'a': 'Root',
            'a.d.#.e': 'The indexed E',
            'a.*.c': 'The Big C',
            'a.*.c.?.u': 'Join',
            'x.y.?.z': 'Union'
        }

        # should find simple nested properties
        self.assertEqual(get_label('a', labels), 'Root')
        self.assertEqual(get_label('@.a', labels), 'Root')
        self.assertEqual(get_label('@.a', content='path'), 'a')

        self.assertEqual(get_label('a.b'), 'A / B')
        self.assertEqual(get_label('a.b', single=True), 'B')
        self.assertEqual(get_label('a.b', content='path', single=True), 'b')
        self.assertEqual(get_label('a.b', content='path', joiner=':'), 'a:b')

        # should detect array properties
        self.assertEqual(get_label('a.d.#.e', labels), 'Root / D / # / The indexed E')
        self.assertEqual(get_label('a.d.#.e', labels, single=True), 'The indexed E')
        self.assertEqual(get_label('a.d.#.e', labels, joiner=' : '), 'Root : D : # : The indexed E')

        # should detect map properties
        self.assertEqual(get_label('a.x.c', labels), 'Root / X / The Big C')
        self.assertEqual(get_label('a.x_x.c', labels), 'Root / X x / The Big C')
        self.assertEqual(get_label('a.x__1_x.c', labels), 'Root / X 1 x / The Big C')
        self.assertEqual(get_label('a.x__1._x.c', labels), 'Root / X 1 / X / C')

        self.assertEqual(get_label('a.x.c.z', labels), 'Root / X / The Big C / Z')
        self.assertEqual(get_label('a.x_x.c.z', labels), 'Root / X x / The Big C / Z')
        self.assertEqual(get_label('a.x__1_x.c.z', labels), 'Root / X 1 x / The Big C / Z')
        self.assertEqual(get_label('a.x__1_x.c.z', labels, joiner=' - '), 'Root - X 1 x - The Big C - Z')

        # should detect union properties
        self.assertEqual(get_label('a.x.c.u', labels), 'Root / X / The Big C / Join')
        self.assertEqual(get_label('a.x_x.c.u', labels), 'Root / X x / The Big C / Join')
        self.assertEqual(get_label('a.x__1_x.c.u', labels), 'Root / X 1 x / The Big C / Join')
        self.assertEqual(get_label('a.x__1._x.c.u', labels), 'Root / X 1 / X / C / U')

        self.assertEqual(get_label('x.y.z', labels), 'X / Y / Union')
        self.assertEqual(get_label('x.y.a.z', labels), 'X / Y / A / Z')

    def test__endpoints(self):
        self.assertEqual(reverse('submission-xlsx'), '/submissions/xlsx/')
        self.assertEqual(reverse('submission-csv'), '/submissions/csv/')

        self.assertEqual(reverse('entity-xlsx'), '/entities/xlsx/')
        self.assertEqual(reverse('entity-csv'), '/entities/csv/')


@tag('nonparallel')
@override_settings(MULTITENANCY=False)
class ExporterViewsTest(TransactionTestCase):

    def setUp(self):
        super(ExporterViewsTest, self).setUp()

        username = 'test'
        email = 'test@example.com'
        password = 'testtest'
        self.user = get_user_model().objects.create_user(username, email, password)
        self.assertTrue(self.client.login(username=username, password=password))

        with open(os.path.join(here, 'files/export.avsc'), 'rb') as in_file:
            self.EXAMPLE_SCHEMA = json.load(in_file)

        with open(os.path.join(here, 'files/export.json'), 'rb') as in_file:
            self.EXAMPLE_PAYLOAD = json.load(in_file)

        self.helper__create_project(1)

        self.assertEqual(models.Project.objects.count(), 1)
        self.assertEqual(models.Submission.objects.count(), 1)
        self.assertEqual(models.Entity.objects.count(), 1)

        self.assertEqual(models.ExportTask.objects.count(), 0)

    def tearDown(self):
        self.client.logout()
        super(ExporterViewsTest, self).tearDown()

    def helper__create_project(self, index):
        project = models.Project.objects.create(
            name=f'project_{index}',
        )

        # create artifacts for the AVRO schema
        artifacts_id = str(project.pk)
        upsert_project_with_avro_schemas(
            project_id=artifacts_id,
            avro_schemas=[{
                'id': artifacts_id,
                'name': f'export_{index}',
                'definition': self.EXAMPLE_SCHEMA,
            }],
        )
        submission = models.Submission.objects.create(
            payload=dict(self.EXAMPLE_PAYLOAD),
            mappingset=models.MappingSet.objects.get(pk=artifacts_id),
        )
        # extract entities
        run_extraction(submission)

    # -----------------------------
    # GENERATE FILES
    # -----------------------------

    def test__generate__csv(self):
        kwargs = {
            'labels': EXAMPLE_LABELS,
            'file_format': CSV_FORMAT,
            'offset': 0,
            'limit': 1,
        }
        # without paths (includes: ``aether_extractor_enrichment``)
        data = models.Submission.objects.annotate(exporter_data=F('payload')).values('id', 'exporter_data')

        with tempfile.TemporaryDirectory() as temp_dir:
            _, zip_path = helper__generate_file(temp_dir, data, paths=[], **kwargs)
            zip_file = zipfile.ZipFile(zip_path, 'r')
            self.assertEqual(zip_file.namelist(),
                             ['export.csv', 'export.1.csv', 'export.2.csv', 'export.3.csv', 'export.4.csv'])

        # with the whole paths list (there are 3 arrays with data, ``iterate_none`` is empty)
        data = models.Submission.objects.annotate(exporter_data=F('payload')).values('id', 'exporter_data')
        with tempfile.TemporaryDirectory() as temp_dir:
            _, zip_path = helper__generate_file(temp_dir, data, paths=EXAMPLE_PATHS, **kwargs)
            zip_file = zipfile.ZipFile(zip_path, 'r')
            self.assertEqual(zip_file.namelist(),
                             ['export.csv', 'export.1.csv', 'export.2.csv', 'export.3.csv'])

        # without `iterate_one` in paths
        paths = [path for path in EXAMPLE_PATHS if not path.startswith('iterate_one')]
        with tempfile.TemporaryDirectory() as temp_dir:
            _, zip_path = helper__generate_file(temp_dir, data, paths=paths, **kwargs)
            zip_file = zipfile.ZipFile(zip_path, 'r')
            self.assertEqual(zip_file.namelist(),
                             ['export.csv', 'export.1.csv', 'export.2.csv'])

        # with `flatten` option should generate only one file
        with tempfile.TemporaryDirectory() as temp_dir:
            _, zip_path = helper__generate_file(
                temp_dir,
                data,
                paths=[],
                options={
                    'header_content': 'paths',
                    'header_separator': '*',
                    'header_shorten': 'no',
                    'data_format': 'flatten',
                },
                **kwargs,
            )
            zip_file = zipfile.ZipFile(zip_path, 'r')
            self.assertEqual(zip_file.namelist(), ['export.csv'])

    def test__generate__xlsx__split(self):
        _id = str(models.Submission.objects.first().pk)
        data = models.Submission.objects.annotate(exporter_data=F('payload')).values('id', 'exporter_data')

        with tempfile.TemporaryDirectory() as temp_dir:
            _, xlsx_path = helper__generate_file(
                temp_dir,
                data,
                paths=EXAMPLE_PATHS,
                labels=EXAMPLE_LABELS,
                file_format=XLSX_FORMAT,
                offset=0,
                limit=1,
                options={
                    'header_content': 'both',  # includes paths and labels
                    'header_separator': '—',
                    'header_shorten': 'no',
                    'data_format': 'split',
                },
            )
            wb = load_workbook(filename=xlsx_path, read_only=True)

        # check workbook content
        ws = wb['0']  # root content

        # check headers: paths
        self.assertEqual(ws['A1'].value, '@')
        self.assertEqual(ws['B1'].value, '@id')
        self.assertEqual(ws['C1'].value, 'country')
        self.assertEqual(ws['D1'].value, 'region')
        self.assertEqual(ws['E1'].value, 'name')
        self.assertEqual(ws['F1'].value, 'location—latitude')
        self.assertEqual(ws['G1'].value, 'location—longitude')
        self.assertEqual(ws['H1'].value, 'location—altitude')
        self.assertEqual(ws['I1'].value, 'location—accuracy')
        self.assertEqual(ws['J1'].value, 'image')
        self.assertEqual(ws['K1'].value, 'number')
        self.assertEqual(ws['L1'].value, 'number2')
        self.assertEqual(ws['M1'].value, 'date')
        self.assertEqual(ws['N1'].value, 'datetime')
        self.assertEqual(ws['O1'].value, 'option')
        self.assertEqual(ws['P1'].value, 'option_a—choice_a')
        self.assertEqual(ws['Q1'].value, 'id')

        # check headers: labels
        self.assertEqual(ws['A2'].value, '@')
        self.assertEqual(ws['B2'].value, '@id')
        self.assertEqual(ws['C2'].value, 'Country')
        self.assertEqual(ws['D2'].value, 'Region')
        self.assertEqual(ws['E2'].value, 'What is your name?')
        self.assertEqual(ws['F2'].value, 'Collect your GPS coordinates — latitude')
        self.assertEqual(ws['G2'].value, 'Collect your GPS coordinates — longitude')
        self.assertEqual(ws['H2'].value, 'Collect your GPS coordinates — altitude')
        self.assertEqual(ws['I2'].value, 'Collect your GPS coordinates — accuracy')
        self.assertEqual(ws['J2'].value, 'Take a picture')
        self.assertEqual(ws['K2'].value, 'How many?')
        self.assertEqual(ws['L2'].value, 'Percentage')
        self.assertEqual(ws['M2'].value, 'When?')
        self.assertEqual(ws['N2'].value, 'At?')
        self.assertEqual(ws['O2'].value, 'Choice (A/B)')
        self.assertEqual(ws['P2'].value, 'Option A — Choice A')
        self.assertEqual(ws['Q2'].value, 'ID')

        # check rows
        self.assertEqual(ws['A3'].value, 1)
        self.assertEqual(ws['B3'].value, _id)
        self.assertEqual(ws['C3'].value, 'CM')
        self.assertEqual(ws['D3'].value, None)
        self.assertEqual(ws['E3'].value, 'Name')
        self.assertEqual(ws['F3'].value, 52.52469543)
        self.assertEqual(ws['G3'].value, 13.39282687)
        self.assertEqual(ws['H3'].value, 108)
        self.assertEqual(ws['I3'].value, 22)
        self.assertEqual(ws['J3'].value, None)
        self.assertEqual(ws['K3'].value, 3)
        self.assertEqual(ws['L3'].value, 3.56)
        self.assertEqual(ws['M3'].value, '2017-07-14T00:00:00')
        self.assertEqual(ws['N3'].value, '2017-07-14T16:38:47.151000+02:00')
        self.assertEqual(ws['O3'].value, 'a')
        self.assertEqual(ws['P3'].value, 'A')
        self.assertEqual(ws['Q3'].value, '6b90cfb6-0ee6-4035-94bc-fb7f3e56d790')

        ws1 = wb['1']  # first array content

        # check headers: paths
        self.assertEqual(ws1['A1'].value, '@')
        self.assertEqual(ws1['B1'].value, '@id')
        self.assertEqual(ws1['C1'].value, 'lang—#')
        self.assertEqual(ws1['D1'].value, 'lang—#—')

        # check headers: labels
        self.assertEqual(ws1['A2'].value, '@')
        self.assertEqual(ws1['B2'].value, '@id')
        self.assertEqual(ws1['C2'].value, 'Spoken languages — #')
        self.assertEqual(ws1['D2'].value, 'Spoken languages — # — ')

        # check rows
        self.assertEqual(ws1['A3'].value, 1)
        self.assertEqual(ws1['B3'].value, _id)
        self.assertEqual(ws1['C3'].value, 1)
        self.assertEqual(ws1['D3'].value, 'EN')

        self.assertEqual(ws1['A4'].value, 1)
        self.assertEqual(ws1['B4'].value, _id)
        self.assertEqual(ws1['C4'].value, 2)
        self.assertEqual(ws1['D4'].value, 'FR')

        ws2 = wb['2']  # second array content

        # check headers: paths
        self.assertEqual(ws2['A1'].value, '@')
        self.assertEqual(ws2['B1'].value, '@id')
        self.assertEqual(ws2['C1'].value, 'iterate—#')
        self.assertEqual(ws2['D1'].value, 'iterate—#—index')
        self.assertEqual(ws2['E1'].value, 'iterate—#—value')

        # check headers: labels
        self.assertEqual(ws2['A2'].value, '@')
        self.assertEqual(ws2['B2'].value, '@id')
        self.assertEqual(ws2['C2'].value, 'Indicate loop elements — #')
        self.assertEqual(ws2['D2'].value, 'Indicate loop elements — # — Index')
        self.assertEqual(ws2['E2'].value, 'Indicate loop elements — # — Value')

        # check rows
        self.assertEqual(ws2['A3'].value, 1)
        self.assertEqual(ws2['B3'].value, _id)
        self.assertEqual(ws2['C3'].value, 1)
        self.assertEqual(ws2['D3'].value, 1)
        self.assertEqual(ws2['E3'].value, 'One')

        self.assertEqual(ws2['A4'].value, 1)
        self.assertEqual(ws2['B4'].value, _id)
        self.assertEqual(ws2['C4'].value, 2)
        self.assertEqual(ws2['D4'].value, 2)
        self.assertEqual(ws2['E4'].value, 'Two')

        self.assertEqual(ws2['A5'].value, 1)
        self.assertEqual(ws2['B5'].value, _id)
        self.assertEqual(ws2['C5'].value, 3)
        self.assertEqual(ws2['D5'].value, 3)
        self.assertEqual(ws2['E5'].value, 'Three')

        ws3 = wb['3']  # third array content

        # check headers: paths
        self.assertEqual(ws3['A1'].value, '@')
        self.assertEqual(ws3['B1'].value, '@id')
        self.assertEqual(ws3['C1'].value, 'iterate_one—#')
        self.assertEqual(ws3['D1'].value, 'iterate_one—#—item')

        # check headers: labels
        self.assertEqual(ws3['A2'].value, '@')
        self.assertEqual(ws3['B2'].value, '@id')
        self.assertEqual(ws3['C2'].value, 'Indicate one — #')
        self.assertEqual(ws3['D2'].value, 'Indicate one — # — Item')

        # check rows
        self.assertEqual(ws3['A3'].value, 1)
        self.assertEqual(ws3['B3'].value, _id)
        self.assertEqual(ws3['C3'].value, 1)
        self.assertEqual(ws3['D3'].value, 'one')

    def test__generate__xlsx__flatten(self):
        _id = str(models.Submission.objects.first().pk)
        data = models.Submission.objects.annotate(exporter_data=F('payload')).values('id', 'exporter_data')

        with tempfile.TemporaryDirectory() as temp_dir:
            _, xlsx_path = helper__generate_file(
                temp_dir,
                data,
                paths=EXAMPLE_PATHS,
                labels=EXAMPLE_LABELS,
                file_format=XLSX_FORMAT,
                offset=0,
                limit=1,
                options={
                    'header_content': 'paths',
                    'header_separator': '—',
                    'header_shorten': 'no',
                    'data_format': 'flatten',
                },
            )
            wb = load_workbook(filename=xlsx_path, read_only=True)

        # check workbook content
        ws = wb['0']  # root content

        # check headers: paths
        self.assertEqual(ws['A1'].value, '@')
        self.assertEqual(ws['B1'].value, '@id')
        self.assertEqual(ws['C1'].value, 'country')
        self.assertEqual(ws['D1'].value, 'region')
        self.assertEqual(ws['E1'].value, 'name')
        self.assertEqual(ws['F1'].value, 'location—latitude')
        self.assertEqual(ws['G1'].value, 'location—longitude')
        self.assertEqual(ws['H1'].value, 'location—altitude')
        self.assertEqual(ws['I1'].value, 'location—accuracy')
        self.assertEqual(ws['J1'].value, 'image')
        self.assertEqual(ws['K1'].value, 'number')
        self.assertEqual(ws['L1'].value, 'number2')
        self.assertEqual(ws['M1'].value, 'date')
        self.assertEqual(ws['N1'].value, 'datetime')
        self.assertEqual(ws['O1'].value, 'option')
        self.assertEqual(ws['P1'].value, 'option_a—choice_a')
        self.assertEqual(ws['Q1'].value, 'lang—1')
        self.assertEqual(ws['R1'].value, 'lang—2')
        self.assertEqual(ws['S1'].value, 'iterate—1—index')
        self.assertEqual(ws['T1'].value, 'iterate—1—value')
        self.assertEqual(ws['U1'].value, 'iterate—2—index')
        self.assertEqual(ws['V1'].value, 'iterate—2—value')
        self.assertEqual(ws['W1'].value, 'iterate—3—index')
        self.assertEqual(ws['X1'].value, 'iterate—3—value')
        self.assertEqual(ws['Y1'].value, 'iterate_one—1—item')
        self.assertEqual(ws['Z1'].value, 'id')

        # check rows
        self.assertEqual(ws['A2'].value, 1)
        self.assertEqual(ws['B2'].value, _id)
        self.assertEqual(ws['C2'].value, 'CM')
        self.assertEqual(ws['D2'].value, None)
        self.assertEqual(ws['E2'].value, 'Name')
        self.assertEqual(ws['F2'].value, 52.52469543)
        self.assertEqual(ws['G2'].value, 13.39282687)
        self.assertEqual(ws['H2'].value, 108)
        self.assertEqual(ws['I2'].value, 22)
        self.assertEqual(ws['J2'].value, None)
        self.assertEqual(ws['K2'].value, 3)
        self.assertEqual(ws['L2'].value, 3.56)
        self.assertEqual(ws['M2'].value, '2017-07-14T00:00:00')
        self.assertEqual(ws['N2'].value, '2017-07-14T16:38:47.151000+02:00')
        self.assertEqual(ws['O2'].value, 'a')
        self.assertEqual(ws['P2'].value, 'A')
        self.assertEqual(ws['Q2'].value, 'EN')
        self.assertEqual(ws['R2'].value, 'FR')
        self.assertEqual(ws['S2'].value, 1)
        self.assertEqual(ws['T2'].value, 'One')
        self.assertEqual(ws['U2'].value, 2)
        self.assertEqual(ws['V2'].value, 'Two')
        self.assertEqual(ws['W2'].value, 3)
        self.assertEqual(ws['X2'].value, 'Three')
        self.assertEqual(ws['Y2'].value, 'one')
        self.assertEqual(ws['Z2'].value, '6b90cfb6-0ee6-4035-94bc-fb7f3e56d790')

    @mock.patch('aether.kernel.api.exporter.RECORDS_PAGE_SIZE', 1)
    def test__generate__xlsx__paginate(self):

        submission_1 = models.Submission.objects.first()
        submission_2 = models.Submission.objects.create(
            payload=dict(self.EXAMPLE_PAYLOAD),
            mappingset=submission_1.mappingset,
        )
        submission_3 = models.Submission.objects.create(
            payload=dict(self.EXAMPLE_PAYLOAD),
            mappingset=submission_1.mappingset,
        )

        data = models.Submission.objects.annotate(exporter_data=F('payload')).values('id', 'exporter_data')
        with tempfile.TemporaryDirectory() as temp_dir:
            _, xlsx_path = helper__generate_file(
                temp_dir,
                data,
                paths=EXAMPLE_PATHS,
                labels=EXAMPLE_LABELS,
                file_format=XLSX_FORMAT,
                offset=0,
                limit=2,
                options={
                    'header_content': 'paths',
                    'header_separator': '*',
                    'header_shorten': '—',
                    'data_format': 'flatten',
                },
            )
            wb = load_workbook(filename=xlsx_path, read_only=True)

        # check workbook content
        ws = wb['0']  # root content

        # check headers: paths
        self.assertEqual(ws['A1'].value, '@')
        self.assertEqual(ws['B1'].value, '@id')

        # check entries (ordered by `modified` DESC)
        self.assertEqual(ws['A2'].value, 1)
        self.assertEqual(ws['B2'].value, str(submission_3.pk))

        self.assertEqual(ws['A3'].value, 2)
        self.assertEqual(ws['B3'].value, str(submission_2.pk))

        self.assertIsNone(ws['A4'].value)  # limit is 2

    # -----------------------------
    # VIEWS
    # -----------------------------

    def test__exporttask_view(self):
        task = models.ExportTask.objects.create(
            name='test',
            project=models.Project.objects.first(),
        )
        task_file = models.ExportTaskFile.objects.create(
            task=task,
            file=SimpleUploadedFile('a.txt', b'123')
        )
        task_url = reverse('exporttask-detail', kwargs={'pk': task.pk})

        response = self.client.get(task_url)
        self.assertEqual(response.status_code, 200)
        data = response.json()

        self.assertEqual(data['name'], 'test')
        self.assertEqual(len(data['files']), 1)

        self.assertEqual(data['files'][0]['md5sum'], task_file.md5sum)
        self.assertEqual(
            data['files'][0]['file_url'],
            f'http://testserver/export-tasks/{task.pk}/file-content/{task_file.pk}/')
        task_file_content = self.client.get(data['files'][0]['file_url'])
        self.assertEqual(task_file_content.getvalue(), b'123')

        task.delete()

        response = self.client.get(task_url)
        self.assertEqual(response.status_code, 404)

    def test__view(self):
        response = self.client.post(reverse('submission-csv'))
        self.assertEqual(response.status_code, 200)

        task_id = response.json()['task']
        task = models.ExportTask.objects.get(pk=task_id)

        self.assertEqual(task.created_by.username, 'test')
        self.assertEqual(task.project.name, 'project_1')
        self.assertEqual(task.status_records, 'DONE')
        self.assertIsNone(task.error_records)
        self.assertIsNone(task.status_attachments)

    def test__empty(self):
        url = reverse('submission-xlsx')
        response = self.client.get(f'{url}?start_at=1')
        self.assertEqual(response.status_code, 200)
        response = self.client.get(f'{url}?start_at=2')
        self.assertEqual(response.status_code, 204)

        response = self.client.get(f'{url}?page=1')
        self.assertEqual(response.status_code, 200)
        response = self.client.get(f'{url}?page=2')
        self.assertEqual(response.status_code, 204)

        response = self.client.post(f'{url}?project=unknown')
        self.assertEqual(response.status_code, 204)

    def test__more_than_one_project(self):
        # create at least 2 more projects
        for i in range(2):
            self.helper__create_project(i + 2)

        self.assertEqual(models.Project.objects.count(), 3)
        self.assertEqual(models.Submission.objects.count(), 3)
        self.assertEqual(models.Entity.objects.count(), 3)

        url = reverse('submission-xlsx')

        response = self.client.post(url)
        self.assertEqual(response.status_code, 400)

        response = self.client.post(f'{url}?project=project_1')
        self.assertEqual(response.status_code, 200)

        response = self.client.post(f'{url}?project=project_2')
        self.assertEqual(response.status_code, 200)

    # -----------------------------
    # ERROR HANDLING
    # -----------------------------

    def test__error__deleted_task(self):
        def my_side_effect(task_id):
            # let's remove the task and execute the real method
            models.ExportTask.objects.filter(pk=task_id).delete()
            execute_records_task(task_id)

        with mock.patch(
            'aether.kernel.api.exporter.execute_records_task',
            side_effect=my_side_effect,
        ):
            response = self.client.get(reverse('submission-xlsx'))

        self.assertEqual(response.status_code, 200)
        self.assertEqual(models.ExportTask.objects.count(), 0)

        task_id = response.json()['task']
        self.assertFalse(models.ExportTask.objects.filter(pk=task_id).exists())

    @mock.patch(
        'aether.kernel.api.exporter.__prepare_xlsx',
        side_effect=OSError('[Errno 2] No such file or directory'),
    )
    def test__xlsx__error(self, *args):
        response = self.client.get(reverse('submission-xlsx'))
        self.assertEqual(response.status_code, 200)

        task_id = response.json()['task']
        task = models.ExportTask.objects.get(pk=task_id)
        self.assertEqual(task.created_by.username, 'test')
        self.assertEqual(task.project.name, 'project_1')
        self.assertEqual(task.status_records, 'ERROR')
        self.assertEqual(task.error_records, '[Errno 2] No such file or directory')
        self.assertEqual(task.files.count(), 0)

        self.assertEqual(task.settings['offset'], 0)
        self.assertEqual(task.settings['limit'], 1)
        self.assertEqual(task.settings['records']['file_format'], 'xlsx')
        self.assertEqual(task.settings['records']['filename'], 'project_1-export')
        self.assertEqual(
            task.settings['records']['export_options'],
            {
                'header_content': 'labels',
                'header_separator': '/',
                'header_shorten': 'no',
                'data_format': 'split',
            })

    @mock.patch(
        'aether.kernel.api.exporter.__generate_csv_files',
        side_effect=OSError('[Errno 2] No such file or directory'),
    )
    def test__csv__error(self, *args):
        for i in range(13):
            models.Submission.objects.create(
                payload=dict({'name': f'Person-{i}'}),
                mappingset=models.MappingSet.objects.first(),
            )

        response = self.client.post(
            reverse('submission-csv'),
            data=json.dumps({
                'paths': ['_id', '_rev'],
                'labels': {'_id': 'id', '_rev': 'rev'},
                'filename': 'submissions',
                'page': 3,
                'page_size': 5,
                'header_content': 'labels and paths',  # not valid, switch to "labels"
                'header_separator': '',  # not valid, switch to "/"
                'header_shorten': 'maybe yes',  # not valid, switch to "no"
                'data_format': 'flattening',  # not valid, switch to "split"
            }),
            content_type='application/json',
        )
        self.assertEqual(response.status_code, 200)
        task_id = response.json()['task']
        task = models.ExportTask.objects.get(pk=task_id)

        self.assertEqual(task.settings['offset'], 10)
        self.assertEqual(task.settings['limit'], 14)  # there was already one submission
        self.assertEqual(task.settings['records']['file_format'], 'csv')
        self.assertEqual(task.settings['records']['filename'], 'submissions')
        self.assertEqual(task.settings['records']['paths'], ['_id', '_rev'])
        self.assertEqual(task.settings['records']['labels'], {'_id': 'id', '_rev': 'rev'})
        self.assertEqual(
            task.settings['records']['export_options'],
            {
                'header_content': 'labels',
                'header_separator': '/',
                'header_shorten': 'no',
                'data_format': 'split',
            })

    @mock.patch(
        'aether.kernel.api.exporter.__generate_csv_files',
        side_effect=OSError('[Errno 2] No such file or directory'),
    )
    def test__csv__error_2(self, *args):
        response = self.client.post(
            reverse('submission-csv'),
            data=json.dumps({
                'header_content': 'paths',
                'header_separator': ':',
                'header_shorten': 'yes',
                'data_format': 'flatten',
                'csv_separator': 'TAB',  # will be replaced with `\t`
            }),
            content_type='application/json',
        )

        self.assertEqual(response.status_code, 200)

        task_id = response.json()['task']
        task = models.ExportTask.objects.get(pk=task_id)

        self.assertEqual(task.created_by.username, 'test')
        self.assertEqual(task.project.name, 'project_1')
        self.assertEqual(task.status_records, 'ERROR')
        self.assertEqual(task.files.count(), 0)

        settings = task.settings
        self.assertEqual(settings['offset'], 0)
        self.assertEqual(settings['limit'], 1)
        self.assertEqual(settings['records']['file_format'], 'csv')
        self.assertEqual(settings['records']['filename'], 'project_1-export')
        self.assertEqual(
            settings['records']['export_options'],
            {
                'header_content': 'paths',
                'header_separator': ':',
                'header_shorten': 'yes',
                'data_format': 'flatten',
            })

    # -----------------------------
    # ATTACHMENTS
    # -----------------------------

    def test__attachments__exclude(self):
        submission = models.Submission.objects.first()
        models.Attachment.objects.create(
            submission=submission,
            attachment_file=SimpleUploadedFile('submission.xml', b'a'),
        )
        models.Attachment.objects.create(
            submission=submission,
            attachment_file=SimpleUploadedFile('audit.csv', b'b'),
        )
        models.Attachment.objects.create(
            submission=submission,
            attachment_file=SimpleUploadedFile('c.txt', b'c'),
        )
        self.assertEqual(models.Attachment.objects.count(), 3)

        response = self.client.post(
            reverse('submission-csv') +
            '?generate_attachments=t&exclude_files=(audit\\.csv|\\.xml)$'
        )
        self.assertEqual(response.status_code, 200)

        self.assertEqual(models.ExportTask.objects.count(), 1)
        task = models.ExportTask.objects.first()

        self.assertEqual(task.status_attachments, 'DONE', task.error_attachments)
        self.assertEqual(task.files.count(), 1)

        # check attachments
        attachments_file = task.files.first()
        with tempfile.NamedTemporaryFile() as fa:
            with open(fa.name, 'wb') as fpa:
                fpa.write(attachments_file.get_content().getvalue())

            _attach_files = zipfile.ZipFile(fa).namelist()

            self.assertEqual(len(_attach_files), 2, _attach_files)
            self.assertIn(f'{submission.pk}/', _attach_files)
            self.assertIn(f'{submission.pk}/c.txt', _attach_files)

    def test__attachments__exclude__all(self):
        submission = models.Submission.objects.first()
        models.Attachment.objects.create(
            submission=submission,
            attachment_file=SimpleUploadedFile('submission.xml', b'a'),
        )
        models.Attachment.objects.create(
            submission=submission,
            attachment_file=SimpleUploadedFile('audit.csv', b'b'),
        )
        self.assertEqual(models.Attachment.objects.count(), 2)

        response = self.client.post(
            reverse('submission-xlsx') +
            '?generate_attachments=t&exclude_files=(audit\\.csv$|\\.xml$)'
        )
        self.assertEqual(response.status_code, 200)

        self.assertEqual(models.ExportTask.objects.count(), 1)
        task = models.ExportTask.objects.first()

        self.assertEqual(task.status_attachments, 'ERROR', task.error_attachments)
        self.assertEqual(task.error_attachments, 'No attachments found!')
        self.assertEqual(task.files.count(), 0)

    def test__attachments__empty(self):
        models.Attachment.objects.all().delete()
        response = self.client.post(reverse('submission-csv') + '?generate_attachments=t')
        self.assertEqual(response.status_code, 204)

        self.assertEqual(models.ExportTask.objects.count(), 0)

    def test__attachments__ok(self):
        submission = models.Submission.objects.first()
        models.Attachment.objects.create(
            submission=submission,
            attachment_file=SimpleUploadedFile('a.txt', b'a123'),
        )
        entity_1 = submission.entities.first()

        # new submission with 2 attachments
        submission.pk = None
        submission.payload = dict(self.EXAMPLE_PAYLOAD)
        submission.save()
        self.assertEqual(models.Submission.objects.count(), 2)

        models.Attachment.objects.create(
            submission=submission,
            attachment_file=SimpleUploadedFile('b.txt', b'b123'),
        )
        models.Attachment.objects.create(
            submission=submission,
            attachment_file=SimpleUploadedFile('c.txt', b'c123'),
        )
        self.assertEqual(models.Attachment.objects.count(), 3)

        run_extraction(submission)
        self.assertEqual(models.Entity.objects.count(), 2)
        entity_2 = submission.entities.first()

        # new submission without attachments
        submission.pk = None
        submission.payload = dict(self.EXAMPLE_PAYLOAD)
        submission.save()
        self.assertEqual(models.Submission.objects.count(), 3)
        run_extraction(submission)
        self.assertEqual(models.Entity.objects.count(), 3)

        response = self.client.post(
            reverse('entity-csv') + '?generate_records=t&generate_attachments=t'
        )
        self.assertEqual(response.status_code, 200)

        self.assertEqual(models.ExportTask.objects.count(), 1)
        task = models.ExportTask.objects.first()

        self.assertEqual(task.created_by.username, 'test')
        self.assertEqual(task.name, 'project_1-export')
        self.assertEqual(task.project.name, 'project_1')
        self.assertEqual(task.status_records, 'DONE', task.error_records)
        self.assertIsNone(task.error_records)
        self.assertEqual(task.status_attachments, 'DONE', task.error_attachments)
        self.assertIsNone(task.error_attachments)
        self.assertEqual(task.files.count(), 2)
        self.assertIsNone(task.revision)

        # export file
        export_file = task.files.first()
        self.assertIn('project_1-export-', export_file.name)
        self.assertIsNone(export_file.revision)

        with tempfile.NamedTemporaryFile() as fe:
            with open(fe.name, 'wb') as fpe:
                fpe.write(export_file.get_content().getvalue())

            _csv_files = zipfile.ZipFile(fe).namelist()

            self.assertEqual(len(_csv_files), 4, _csv_files)
            self.assertIn('project_1-export.csv', _csv_files)
            self.assertIn('project_1-export.1.csv', _csv_files)
            self.assertIn('project_1-export.2.csv', _csv_files)
            self.assertIn('project_1-export.3.csv', _csv_files)

        # attachments
        attachments_file = task.files.last()
        self.assertIn('project_1-export-attachments-', attachments_file.name)
        self.assertIsNone(attachments_file.revision)

        with tempfile.NamedTemporaryFile() as fa:
            with open(fa.name, 'wb') as fpa:
                fpa.write(attachments_file.get_content().getvalue())

            _attach_files = zipfile.ZipFile(fa).namelist()

            self.assertEqual(len(_attach_files), 5, _attach_files)
            self.assertIn(f'{entity_1.pk}/', _attach_files)
            self.assertIn(f'{entity_1.pk}/a.txt', _attach_files)

            self.assertIn(f'{entity_2.pk}/', _attach_files)
            self.assertIn(f'{entity_2.pk}/b.txt', _attach_files)
            self.assertIn(f'{entity_2.pk}/c.txt', _attach_files)

    def test__attachments__deleted_task(self):
        def my_side_effect(task_id):
            # let's remove the task and execute the real method
            models.ExportTask.objects.filter(pk=task_id).delete()
            execute_attachments_task(task_id)

        models.Attachment.objects.create(
            submission=models.Submission.objects.first(),
            attachment_file=SimpleUploadedFile('a.txt', b'123'),
        )

        with mock.patch(
            'aether.kernel.api.exporter.execute_attachments_task',
            side_effect=my_side_effect,
        ):
            response = self.client.post(reverse('submission-csv') + '?generate_attachments=t')

        self.assertEqual(response.status_code, 200)
        self.assertEqual(models.ExportTask.objects.count(), 0)

    @override_settings(EXPORT_NUM_CHUNKS=1)  # creates 3 processes
    def test__attachments__error(self, *args):
        def my_side_effect(*args, **kwargs):
            if not kwargs['url'].endswith('/b.txt'):
                safe_sleep()  # wait a little bit
                return requests.request(*args, **kwargs)  # real method
            else:
                # there is going to be an unexpected error while fetching file "b.txt"
                raise RuntimeError('Being evil')

        models.Attachment.objects.create(
            submission=models.Submission.objects.first(),
            attachment_file=SimpleUploadedFile('a.txt', b'123'),
        )
        models.Attachment.objects.create(
            submission=models.Submission.objects.first(),
            attachment_file=SimpleUploadedFile('b.txt', b'123'),
        )
        models.Attachment.objects.create(
            submission=models.Submission.objects.first(),
            attachment_file=SimpleUploadedFile('c.txt', b'123'),
        )

        with mock.patch('aether.sdk.utils.request',
                        side_effect=my_side_effect):
            response = self.client.post(reverse('submission-csv') + '?generate_attachments=t')

        self.assertEqual(response.status_code, 200)

        task_id = response.json()['task']
        task = models.ExportTask.objects.get(pk=task_id)

        self.assertEqual(task.created_by.username, 'test')
        self.assertEqual(task.name, 'project_1-export')
        self.assertEqual(task.project.name, 'project_1')
        self.assertIsNone(task.status_records)
        self.assertIsNone(task.error_records)
        self.assertEqual(task.status_attachments, 'ERROR')
        self.assertEqual(task.error_attachments, 'Being evil')
        self.assertEqual(task.files.count(), 0)
        self.assertIsNone(task.revision)

    @mock.patch(
        'shutil.make_archive',
        side_effect=RuntimeError('Zip too big!!!'),
    )
    def test__attachments__error__zipping(self, mock_req):
        models.Attachment.objects.create(
            submission=models.Submission.objects.first(),
            attachment_file=SimpleUploadedFile('a.txt', b'123'),
        )

        response = self.client.post(reverse('submission-csv') + '?generate_attachments=t')
        self.assertEqual(response.status_code, 200)

        task_id = response.json()['task']
        task = models.ExportTask.objects.get(pk=task_id)

        self.assertEqual(task.created_by.username, 'test')
        self.assertEqual(task.name, 'project_1-export')
        self.assertEqual(task.project.name, 'project_1')
        self.assertIsNone(task.status_records)
        self.assertIsNone(task.error_records)
        self.assertEqual(task.status_attachments, 'ERROR')
        self.assertEqual(task.error_attachments, 'Zip too big!!!')
        self.assertEqual(task.files.count(), 0)
        self.assertIsNone(task.revision)
